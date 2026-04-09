"""结果导出模块 - 导出JSON、Markdown或XLS格式"""
import json
import re
from pathlib import Path
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXLS = True
except ImportError:
    HAS_OPENPYXLS = False


def export_to_json(results: List[Dict[str, Any]], output_path: str) -> None:
    """导出结果为JSON格式"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def export_to_markdown(results: List[Dict[str, Any]], output_path: str) -> None:
    """导出结果为Markdown格式"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# 论文关键词知识点提取结果\n\n")

        for paper_result in results:
            f.write(f"## {paper_result['paper']}\n\n")
            f.write(f"**文件路径**: `{paper_result['path']}`\n\n")

            keywords_found = paper_result.get('keywords_found', {})
            if not keywords_found:
                f.write("*未找到关键词*\n\n")
                continue

            for keyword, sentences in keywords_found.items():
                f.write(f"### 关键词: {keyword}\n\n")
                for i, sentence in enumerate(sentences, 1):
                    f.write(f"{i}. {sentence}\n")
                f.write("\n")

            f.write("---\n\n")


def _default_polish(sentences: List[str]) -> str:
    """默认整理：去重 + 连接"""
    unique_sentences = []
    seen = set()
    for s in sentences:
        s = s.strip()
        if s and s not in seen:
            seen.add(s)
            unique_sentences.append(s)
    content = "。".join(unique_sentences)
    if content and content[-1] not in "。.!?;":
        content += "。"
    return content


def _llm_polish(sentences: List[str], keyword: str, paper_title: str, llm_client) -> str:
    """使用 LLM 整理润色多个句子"""
    if not sentences:
        return ""
    unique_sentences = []
    seen = set()
    for s in sentences:
        s = s.strip()
        if s and s not in seen:
            seen.add(s)
            unique_sentences.append(s)
    sentences_to_process = unique_sentences[:10]
    sentences_text = "\n".join([f"{i+1}. {s}" for i, s in enumerate(sentences_to_process)])

    prompt = f"""论文标题：{paper_title}
关键词：{keyword}

以下是该关键词在论文中出现的相关句子，请进行整理润色：
{sentences_text}

要求：
1. 去除重复内容，保留核心信息
2. 修正语序不通、错别字等错误
3. 将相关联的内容合并，使表述更简洁流畅
4. 保留关键的技术术语和数据
5. 只返回润色后的内容，不要解释

润色结果："""

    try:
        response = llm_client.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        return response.content[0].text.strip()
    except Exception:
        return _default_polish(sentences_to_process)


def _clean_sentence(sentence: str) -> str:
    """
    清理句子：合并断行、去除多余空白，写成一行

    Args:
        sentence: 原始句子（可能有换行、断行）

    Returns:
        清理后的一行句子
    """
    if not sentence:
        return ""

    # 合并换行符和多余空白
    cleaned = re.sub(r'[\n\r\t]+', ' ', sentence)  # 换行、tab换成空格
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)       # 多个空格合并成一个
    cleaned = cleaned.strip()

    # 确保以标点结尾
    if cleaned and cleaned[-1] not in "。.!?;：,":
        cleaned += "。"

    return cleaned


def export_to_xls(
    results: List[Dict[str, Any]],
    output_path: str,
    llm_client=None,
    progress_callback=None
) -> None:
    """
    导出结果为XLS格式，每个句子独立一行

    Args:
        results: 搜索结果列表
        output_path: 输出文件路径
        llm_client: LLM 客户端实例（可选，为 None 时使用简单去重）
        progress_callback: 进度回调函数 (current, total, message)
    """
    if not HAS_OPENPYXLS:
        raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "关键词知识点"

    # 表头
    headers = ["论文标题", "关键词", "PDF路径", "原始句子", "整理后内容"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 计算句子总数（用于进度）
    total_items = 0
    for paper_result in results:
        keywords_found = paper_result.get("keywords_found", {})
        for sentences in keywords_found.values():
            total_items += len(sentences)

    # 数据行
    row = 2
    current_item = 0

    for paper_result in results:
        paper_title = paper_result.get("paper", "未知论文")
        keywords_found = paper_result.get("keywords_found", {})

        if not keywords_found:
            continue

        for keyword, sentences in keywords_found.items():
            # 去重
            unique_sentences = []
            seen = set()
            for s in sentences:
                s = s.strip()
                if s and s not in seen:
                    seen.add(s)
                    unique_sentences.append(s)

            for sentence in unique_sentences:
                current_item += 1
                if progress_callback:
                    progress_callback(current_item, total_items, f"处理句子...")

                # 清理句子：合并断行、去除多余空白
                cleaned = _clean_sentence(sentence)

                ws.cell(row=row, column=1, value=paper_title)
                ws.cell(row=row, column=2, value=keyword)
                ws.cell(row=row, column=3, value=paper_result.get("path", ""))
                ws.cell(row=row, column=4, value=sentence)      # 原始句子
                ws.cell(row=row, column=5, value=cleaned)       # 整理后一行
                ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
                row += 1

    # 设置列宽
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50

    wb.save(output_path)
